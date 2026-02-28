import openpyxl
import json

file_path = r"D:\SAFFET SABANCI APARTMANI\2025\SAFFET SABANCI APRT. 2025.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print("=== SHEET NAMES ===")
for name in wb.sheetnames:
    print(f"  - {name}")

print("\n\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")
    
    has_data = False
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                row_data.append(f"[C{cell.column}:{val}]")
        if row_data:
            has_data = True
            print(f"Row {row[0].row}: " + "  ".join(row_data))
    
    if not has_data:
        print("  (empty sheet)")

wb.close()
print("\n\nDONE")
